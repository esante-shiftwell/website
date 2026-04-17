from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
OTHER_SOURCES = DOCS / "other_sources"
WORKBOOK_PATH = DOCS / "Fatigue Index_scoring_system_15.xlsm"


def cell_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def extract_pdf_text(pdf_path: Path) -> Path:
    out_path = pdf_path.with_suffix(".txt")
    reader = PdfReader(str(pdf_path))
    text = "\n\n".join((page.extract_text() or "") for page in reader.pages)
    out_path.write_text(text, encoding="utf-8")
    return out_path


def export_workbook_snapshot(workbook_path: Path) -> tuple[Path, Path]:
    workbook = load_workbook(workbook_path, data_only=False, keep_vba=True)

    snapshot = {
        "sheet_names": workbook.sheetnames,
        "parameter_cells": {},
        "graphique_headers": {},
    }

    ws = workbook["Parametres score"]
    for cell in [
        "B6",
        "C9",
        "E6",
        "E8",
        "E9",
        "E11",
        "E12",
        "E13",
        "E14",
        "E18",
        "E20",
        "C24",
        "E24",
        "E26",
    ]:
        snapshot["parameter_cells"][cell] = cell_value(ws[cell].value)

    graph = workbook["Graphiques brut"]
    for row in (4, 5, 6):
        snapshot["graphique_headers"][f"row_{row}"] = {
            col: cell_value(graph[f"{col}{row}"].value)
            for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
        }

    json_path = DOCS / "after" / "workbook-snapshot.json"
    md_path = DOCS / "after" / "workbook-snapshot.md"
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")

    lines = [
        "# Workbook Snapshot",
        "",
        f"Source: `{workbook_path.name}`",
        "",
        "## Sheets",
        "",
    ]
    lines.extend([f"- `{name}`" for name in workbook.sheetnames])
    lines.extend(
        [
            "",
            "## Parameter Cells",
            "",
        ]
    )
    lines.extend(
        [f"- `{cell}`: `{value}`" for cell, value in snapshot["parameter_cells"].items()]
    )
    lines.extend(
        [
            "",
            "## Graphiques Brut Rows",
            "",
        ]
    )
    for row_name, row_values in snapshot["graphique_headers"].items():
        lines.append(f"### `{row_name}`")
        lines.append("")
        for col, value in row_values.items():
            if value is not None:
                lines.append(f"- `{col}`: `{value}`")
        lines.append("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    return json_path, md_path


def main() -> None:
    outputs: list[Path] = []

    for pdf in OTHER_SOURCES.glob("*.pdf"):
        outputs.append(extract_pdf_text(pdf))

    json_path, md_path = export_workbook_snapshot(WORKBOOK_PATH)
    outputs.extend([json_path, md_path])

    print("Generated files:")
    for path in outputs:
        print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()
